#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Envio automático de PDFs de credenciais (SSHD) para colaboradores -
Versão com interface gráfica, e-mail em HTML com logo, e cópia
automática salva na pasta "Enviados" via IMAP.

Pra virar um .exe:
    python -m PyInstaller --onefile --windowed --name EnviarCredenciais enviar_credenciais_gui.py
"""

import csv
import imaplib
import json
import os
import re
import smtplib
import threading
import time
import tkinter as tk
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from email.utils import formatdate, make_msgid
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk

import openpyxl
import pdfplumber
from unidecode import unidecode

CONFIG_FILE = Path.home() / ".enviar_credenciais_config.json"

COLUNA_NOME_PADRAO = "Nome Colaborador"
COLUNA_EMAIL_PADRAO = "E-MAIL"

ASSUNTO_EMAIL = "Seu acesso ao Complexo Hospitalar dos Estivadores"

# Nomes de pasta "Enviados" mais comuns em servidores Zimbra, testados
# nessa ordem até um funcionar.
PASTAS_ENVIADOS_CANDIDATAS = ["Sent", "Enviados", "Sent Items", "INBOX/Sent", "INBOX/Enviados"]

LOGO1_B64 = """iVBORw0KGgoAAAANSUhEUgAAAEgAAABcCAYAAADNqvPKAAAQAElEQVR4Aex8CZhV1ZXuv/eZ7r01D1QVUEwCgoDiACqOqDExdIZuTaJ5rzvdrzu2nS+dmMHuFzOoMemYRH1GVGQegmg0xnQn6cS0ncR0oiKigBRjMVNQUCM13PEM+/3rVBWiEVQuvPd973uHs+4+Zw9rrf3vtddee59SbYq8osiYKIpIgYlMgZQjZUj9pD5S1kSRb6KQ9UiFfGR8PzBhmDNB2GP6882mte0V07z7ebNlx3Nmd8sfTMvh1SaT327C6AApbcLQN3KFgSEvId9Epp/UQxJ5vomiwEQR9QhZfgyZIi+Noi8zyIGsjJANGIfkDRAslivSwG2zSGkfBb8XO3ZvxNpXX8Da9S+iacur2LVnK99fxKvrXsIr617AgbZdiKIcDAowJiJRFm/EPMkIJEP+RpE5ZfMX8ijpKaJBrsVwE42FhIcoKyQgCQ12QLRWrKMiVsojl+/Ehk1/RPPO19DbfxiW48NNRCgEvUiVaiRTCj19h/F60xpsbV5PYHIEKo/IhGzPWwCJByHBQaCcGLAhZFhHBYDyAQ4EK6CY6xQBNKgCMYifJB0iEJQhhVUeftiLnXtex6G2HdB2GkrnEYRpdj4LxwlpLVnkCz3wPIMwSmPPvs3YtXcTLKsAraXzwphSBCQm8T2YhViWyJN6BAlCcY2T/jkFACkKF3ojoSdg5ww7HcEP+tB9pBV79m5FW/tetHfuY6e3wEuyIwRH6YAdD6BkxAdJ8vwgA8uOINa1d/82dPUexMFDO3Do8F5aXQf8qJ9wZEkFWl6Wwn2CGzI1gyQ6CfG1iPsUACQsRBFRLECevkVbnA7owf6Dm/Dyq8/hldd+h9fW/wG96Ta0HNxF8FgeFQheDpCph+NdBgJcOtuJvnR7DNCr65/HS2uexcbN/0Wwt7M8A88NyDMDrYYAkmnuAoYE0e14/N85X3r3zrVOVMOwUIjjB04l2wnQ3r0bq9f+Ghu3/JG+ZD+g++hbDFIlCkd62uF6Nq3GgmOLI2dnYh/ydqkmiHk47Of+lmbUDy+D0b20njYcPLwRa9Y9i1c3/A7pXAenYIggylMZBRj6JVkwouK7VzwHqhTfsQMucArsxpq1v0M23w7L4ajaOeQKR+AlLTiOhXw+zzFVKBQCNhNQRIXjkUK+kIPnOSj4WXgJDcv2oWih0DloO4vD7Xvw2oYX0NbJgYh1IFvxT/GgyTOpiFs0K6L5QNPQFDh6fWg5tIWr0x/YgQzf+2lTnEomQCqZgl8wUCrBaZBE4FsccZerE/3QAIvj/rqOy3aKdRXAJd2EfDfkgxRnZ5I+qB/dPS1o2rwW3d2HEIU5TrcCKaLzNwCboYhLF9E2bmoIg+Lc7+vtwrp1rzCPfiAeSSrHtygKCZaPXC7LToJW5HJ6KdiOltJ3JHH42WwWJakSZHPis9jk6G1oUZr8gL7+boK0HtlCHxiYUquQg1AkOpQjWjIp5vY5bfqwrXkTlSUoMTjRIEMTK1koFABl0NNzBJUVFRxZH77vIwwD5hNQHI8iWNqiVQB1dfU4cqQboM2A3QcHJSa2tWxFP6WQznRj2/YmuK4AY1i3+LtIgAyULtAP7EdndyufhzocUTMhBdt24ecDJLwUOtq7MGJEI2zLZR81XFcCPcW6x7sV66RIJRhWOxxHuvtZUZOGbspXnH6Md7QVwqgcOjoPoINTzbY1LdYMVTzp9Fhp74qJMZxUR0lA8Ll07+QIUhnGNOBKhnh0pczALyhUV4/AGWPPwjnTZqK+djRGNU6gVTgwkYMgYAcZFSsIaA51cJHPsS3zpDyX0Rg3ZgpSbg0um3UNpk09H8lkGX2ZQ+tyMaBPgMjkoK1CvCB0dh0knxBaK6bF3bq45gbpbDd6+zpiBXN5GWGZLuwgCBg9pDJJgjMVE8aeA0uVsVNlGD/ubNQPGwsTJpFwK1grQcdtwxfnrVJ06tXxs1ZJjB11NsaPno4oSsXtGxsm4qxJ58XtBHywNWRax0HmAEgdna1xeBCEoguKut47QG8SR4DS3VAWl15ajuNwxGJlBRwhoL5uLMaMmAwTJeA65VAEIElrmD5tFiaOn05LKGH/kkw5lexy5DKIwaquGIGpZ80kGDOhUMbyUijFuly9aqtGobysjnkep6ooxAFRPmT/ZdkhB6yLiwIZofirSICATK6PK5NGLp+mvyFAkFGjwrEFAclEObX0EISaVqAR0F9HoQPXrsD4MdNw6YXvw/nnXoYzJ5zLqTQVM2dciZnnD1AjrUWsCPAQ+g7bWggDG7ZOoKpyGMGxoPiPDwCdt1GcVpylkQmQpTWLH2JBUXeRABkUGHdkCU5IR1lg7yMqGhEcgSii8iWpMojCmjtrTUtzXA1x0srQIkwtkt5o1NdMxcSxszB5/GUYMewc1FVPQcJphEYNtEUI6EscOl3XsmEpD4r8E05IV9cHOzTQkaJUIOBmNmeyKKgMDPd5UcR08JjEDPpNvMerSIAUHCsFbTzYSLGzFXB0Cg6n0VAKiAg1qJZMu2jwmYm8wuKDfRySMgXFUrmHAuSI7QxzteXAHpRn6xL6twGyCaLFoFLLdkMaFkGifRHNFcEoh6urCFIZwryHyE8OUEHSBGCk8yJGuilEcYMJ+8iXE98qRoUNeEtNsajAAFn+5EILfuhx+iY5BUVeijokYKiDZRJk71L+YENpfBIkmp9Es6EmGtmcDcerRXlFIzek9SgpGY6S0kEqq6eCBCgeyWNFsYfvSm/FTr7RTrAKYRBpG26yEuXVI5Gks06UNiBVNoKb4ZGorBiN0pJhsWUhljuk68mlb0g/ifaG6o8YcyZmzLoaF1x4Fc6fORszL75mgBizXMj8+vpGZLOMT3L0zoZClEwxIb4oA6UIwjHEGm++2UnFHEOSzbkAlPMjlFU04IIZV+G8WbNx/sVX4byZ1+AC0nnnXYELmVdaJaFCnq2Ku4sECPjtSy9h7pLFeGTlcsxduhQPLFlIWoQHFi/EDxYuwrr1TVzlkiSPO3OCJPoKSCoAvSxoYqQT3IKMkBqomQ99LgcKr2/bgYeXLscDSxfi/mWL8MCyJfjBsqWYt2olfrDgUWzdvR1c/AAljXHSV9EAbWvZi4WP/xAPLl6AJU89juVPP0X6EelJrHj6SWxs2gytbLiOB4ur0ICmorRYkdBAztv+SjWhwULD6QUC5bopbN6+E3PnLcLynzyJ5c/8CMue+TFW/OTHWEiAlj21CjtadyETchUbbHuySVEAidACx1OnXJiEAyRd5GgdQ5Tns5dIQGsNrrLcXhAQeTja6aMPwupPiFi8KU8RoDCUScZsZaGiuhYFCyiwFwVt4jRrAkSOhvZsUDBbnFgGOZ3wJusTlr9DIcfU7wEYb9ieRprmn1VAjq18KhpGBdj0L4giiKCEl4SiI4mJK5Ai8eAGYKeHSPF4ZIi4FwE0QSXQmvGMw/W9ghtcCxEsMsyHOR6lsDmrUETMytYOZTjQKgHXLmVqQSl1lKjae7op5j3Vf0tlA0VliRDzCVb8qzhqfJCb879AkEIBi59sIknZ2fBNFCJkBPwGRXx/gyICZFifvQbIL4rlAbABzeBRNAC1wFsvWfLeLv+t9d7hXb9DeVHFhq1DmrvPaNi3AF8rBO+RCuTBAwH4iu1JBQJUgIBqQGNk6em9Ty9AHPE0ChiiDHz0mMwJqRdZDFEPn/s4YYX6maZJGZNHhvlZpgUe9eINez0tSJ1ygLTW9I0DpPj88PIFuOaGObj2Ex/C1R//IN53/Z+dkK75izk4StfPwewbrsNV15P+4jpc/dHrcC3zPnLTDViwfDEii/DQqk4LMoNMTzlAg3zjRNxAmlbTFfSjLXMkpiN06O+WerhMd0t9pj38kCjvHTx7PtTdhiMZLg7uaVU/7sNplSA+KHIUJ0aADKdDTgXI0yG/W8qxbk6HGKIC3wMrgvIsGFshU8ie5gkGnFaAZAh8Ls8FLt2h4pttIc+D+ndLhShAbpDyTIW056JAfgHXdS/JsCGGyECsVRYtGRRK4v3GE19O+j7tAEmEaGuLI6EQBTzDod703YjpHdSWmBKsz1vYxCRfQwwLhAK/AIsDoLiqRWJdtLaAh/cRz8Y182wShME7yDlRsT5R4akos9i741IEcMYcn6Qt69ikozwGn0XxAZBZKVaUVhSDISnEmHAqLpFzKvi8LQ/FXJtz60Tk8DTweOSyrSfELx8eyeGzzfrCz2JqcadPEaf1Pr0AcXAdnkV7J6CEr3EiSkp5oOENkkteNiNEm6mApGLnc/owOuUARXSeQ2T4/FfX34Qlc+fjsfnLsOKRxVj28MKjtJzPP3x4MVYeh6RsxYMLsHzuAC17aAFWzFuM5aS//cu/gaec04fMIOdTDtAg3ziR0R01bARmTD4fU884C+eeeQ5m8JvWsTRz0nTmvT1J2ayzZkDo4qkzcBHp7AlTcc6EszG8uh7Zvuwp8zU4znVaARKZIU8TPfoKlz4jAQuOTLtjyOUU8U5AsoN3uYt3+PXCZuopi+dsEWSalSVLRMRppSIBUlw3dEzgkcKApuy9PCj5AVKuB3CquTw0swmETZBsfnEYIkXgcALiZh4DdagqfY/m/sLm1NIMHMJCOCAk/lWQf/IoWslaJiTvxRClnnxzQ5UKVgmDOQd9/Tl4ls0vHICjwf2YDaVdfm3IQTOKdtjTgJGvpQBLaRKhImjadnAsKfIYIm1Z0GynNLtsM6ohZWiNIQ/t8/mAEDFfUX+lICuaQ6A9ngcZBpKykU37/TCMk8xg3CQp3uOl32P9N1UX3coSpQgzeaQsFyrvw/ENnACw+VnG5rTwC3l4joM4qLM0IRVYwRTveBmpIUJoo/Io7+L4LaXhMCoHo3KbshzKGiKdDxFlfTDchms5wLuShONe+rgl76JAdB9fNwIfvvIa/PnV78cHL70SH5h12SBdjmtnXY5JE8+E/BGUHJWqd6vsm2QrgBYCXnyKgTE8uRzbOALXXTMb76e89198BT5w8aWUdynmzJ6Nj1zzQYyrG40UPypKGzY96btogK69aBYe/vZ3ce/t38B9X/0GfnDntwborm/hgbvvxiUXXYz+dJp9VDG9d01FRRU3UzQheYv8HC4+bzq+d+edlPUdPHgHZd7xTTxA+t5X7sA9d9yJaROn0pLYgH4vbnySPyLvJJsC9AA4vL8ZzZvWomn9C9i57TVs2vACNvF50/oXsWndS+jt6UYykYTr8mAfJ3exm5QF+hzE/s1zFPp45LG5aS02r1uNzRsoa+MLaGr6I7Zuexk7t29EtrcHWuuTE3hMqyI5RHCdAlr2bUL74R1oP7QTnW270dlOatuFjra96OzqoAcxUErFdIzsd/XIlqynSID8ajrgJD196KfRsrsZXZTRFcvcgY6OZrQe2o6WA9vRn+mG0gbFXkUCZBAEPVCqD4mEz68KPXCcLBybxNR28tQvgs3VKOJSX5C/VWTOu79po0pgERpsxcN/v5BBmEvDrSd/+AAAEABJREFU5Q7esQuUl4Pt5GBRJuw0j1R6YHSOFACqOJCKBsjoAmD5yOSO8HAhS+X6j1IhSBM8BdAPMMajtXnxM7UGHwYpZHoiIkisIbfici0ddrgqJlyHwCsUeFqZD/vhM/WZaqtAoCI4SRdSF0VeRQKkqEwCbrIUhvGLRT8DptQcsMiaU8EPfChtw0Q2osjhymwxjdihLAwy/K7VgXR2P7p6d6Ktaxv6Mnt5UtgCqG5E6EZYiGC4iw+5lAvIEX8MS3JRHiAGZAnYiuIUXGPD5ICELkPIzyiRcTlAmqSOEt7jpd9j/T+pnnAZKEr8YyeQZwwCWKwjbBVkBMMwS0CysO2I2IUxaSsghgGO9B7C+qbV+OPq/8R//fHXeG3dC/jd73+F1S8/j41b1tAqu6AZ6IT8gqHZ3icvTWsNoiy6j7QRJh+G/ygIoAtXxmLsI3/2YiHhlUAxD0Ve0pMiWCiUlVbDUtxlWQkYKghDlpxSoHrg1dLajIOHmzn9eglUL5RKIwx7sGv/Rqxe8xt0dNJa6C+SJZxKOgsvZdjXHA607sTqV57HQTpdh77MD7ohKVi3o3s/evvbYHjaDUUhAoRhUEiKQhulJZXw5CuujgulwkmTPumWcUOFVLKKCtXA0NYdO8ncYyyIXQABad61DvsObEa20IGC6caBtm1o2rQ6Bi00fYhMhg6+n9Oul9SHvnQblJWDH/aiactLONSxA7AyKIRH0NK6nUv5euTyXBBcgiky4kERgGjFOYWa6gZOOQdRGFGf4u73DJBSCkoNkTR3MWb0mUDkchxdgH4AhiCJ0tRN0WnKf0m4acur2ND0Mjvfji1bX0OyRAqzkL9Mlb9xNvzAaHEaaU4/h2zCiKsQp5ZRWWxrXo90poPT71kC+xqyObFGn6AWBnWxoWnFUegi6VVhWG0j3x0ACsVeujgGnBa0nLraUaiuakBERxyDA2ErRO781JPL98PioXJ1TRkOHdrPjuXhB1loS0ZYrID13vY29FkRY6lWBGEOlVXlrDVUn53nVDZ8VbBpLRZdXgK1NSNRS12CIIJSrMMWxdy6mMYDbR14TinOnDgF3DvSgoSl0ECp7KAJI8FQqC6vwoGDB+Bwic7mMgTNGqgfW5w8v4UItM/YyXFs7N+/D/X19WQqnRb+QqzPOrLCyX9NVJKqwqQzp6FQIGqQeqxe5K2LbE81RBE7dtbnnTuT7Kh0PL0kH7SqCMlkkvmAV+LxPUR/Xx8trpqrXoH5miR13440DI8rLAaafX39cOVsiRIR8x9oJ+BEPOYoK63CtCnncoqVQWubtSyE/zd8EHvzpzctQKsSNDachamTZzEG4YoWJmDpJJXVnFIBOxqSDJUOYDsakQkgIRPYFUA6+3ak4HoOlDKkCB6/qEaRD8NpC37SlqMPWw1DaWIUzjvnClRWinP2YNsOSXPV0yj2Kp7DkAbxqLpoHD4Rl856P8pK6pHp13DtciTcSvoIh6YfUmkPiUQCQVAgYCFbn5hysqXwbDrpNLI8lLMsDyG/cISBA8cuQ231CMzicUdleQ0UfRH4i1N4FQ+QojZCopisYEiiqnw0Ljz//Zhx7jU0+QZk0w6j4RTCvAPXSSGbzcdWZTMCBvdTEIs4DpWWlSCdTmP48JHo7y8glzX0eVUYNfIsTqlZmHn+JfDcEhhYsLUzpAxEHZyCq3iAGM+C6iHWyOZkSfA1wU5UsxNTMWvGB3DuOZdh+rRLoDkNx4yaCCPnytqDia0uAhsch4AwNCjkIzSOGIsqrk6zLroaF894H6ZOmoWRwyZzqtps7iAKFPkp8hmiwUcmxdynACBD+UIDiVKKyy1JAKPztKwSDK8bxw5OQEPdWNTVjuG0GIUwcCFxi1ZO3DEBQgjQUMoiMx1PyaBgY+L4aaipaGR8MwoNw8aiorwOMr1oS5D4R5GH7XBg2M4MLv3GGPIlmyJvXWR7NlekwXvoUdIhgogYGGXAoz+qiv8TJ8+pRZAXX2RDqyQcAilpIa+gkUCeU8nWJagsb8TEM86HQiksVQoYj+QAsEjkze0FYhIZg3ksGbgHB27g5aR+KeGk2h3TSPFZiAk3DzjWp2ifmUNTiKLooww7KNYw87yrMGrEFFpbgtZk02lLBz0CWM6R9+hXytE4cgKmT72UeXUEpYR+zGVqA4SQP4O3xTzyjvMkHcyOk/8jAImQ45FooeSHRCC4HEMFGCCJcYRklZL24KVh0fdEjL5LU3WYwq+mF5x7GcaMmoSK0nqUpmo5fRowauREXDzzKkybPAOuUwnNbYThNzHQjhBfIlN4knjHWUPp0TpHM+Lik/15K+Rv4SNC2HGxjOMQXQ6UUiQLimav4DL1SAlScpCYpzRYDRxupja0TiDh1NCnTMPZk6/G5bOux5WXfIxO/SOYeuZsVFdM4qpUT1+TYn0NbSkSoDRJUV4shzLkfYiUlAmpuI2SyijuIut3YiAgHa+OlAkNlSs+DJGwFhp6Z9Gb7qF8i7kybd6OpIzFb3sPtX/bwlOWKT04Zcz+X2T0/wF6h1H9E4AkfggCOlo2lOeI8UTAbXrETWMURVxtAkgahiGEosjE7xHLhAbyoqN5wkvyJZUypRSUUuROb8TZSfYxT5FlyEvqKTDoi59DyN8kSnshKZN6cWP+yLMQH+NbnkWGpEJDz3Ehf+Rd+BwvlTJWe9P9tgBZ3D2LAN8PuIKouLPSSvEI07Ztxs3smQLkXeodS0opLtMsB6C1hvASwdJO3oeUkzasSihAR+xQjj5a31CCDIi0sbnxJKtYB2k/9Cw85VlIeArJs8gT3kLyLPUEWCmX9kopKKViWfIuZeCllOLvn976rVnCUPJEgMtzG2HsUEkZSckTaxKmwlzqaYImz0OklIoVED7SRlKboEpdaa+Ukse4jrxHNKGQHwOlXkQrlUL5fmZpa9BCByxV8oWUGhgAaavUwLNS6ig/yR/SRZ6ljci3OOiit+RJeTA4S1z5EsNK0k+lFJ/efGsZ62PJsi3kCzme4rVzB92P/kwf2jvbYNsa7R2HeYaT5RYgx2PPDL9g7o/rSv0hOth6IG4XhD7PjVnXzyOTTUPyDQGITMjTxEJcp7dPjk6DmG8hyMX1O46Rlc30wy/k4ZNHX38vDrdTPr/LC79sJg2fevYzP4yCePMrdSKC3d7RRp5ttLoQ2gL6+nrQxS+8fpCPdW8nH6UMenuPoK1d6kXsn4WAh3M0fyI0gIihJWu6DnC6DxAzfH5a2ccvEftbtyEXdWDbrrXYc+B1dKf3YeO2l9DauR1NzavR1bMX+1o3o5WfmTdtexW2G6Av247DnXvheCEO8tPzlh3r0Z/vwvY9TTjMLxFHsm0wVh4dvQex79AO7D/UjA4eyAvt3LMOu/dvQEfPHmT9w9jf0oTt29ag/0grWg/sgHZ9bG/ZjIzpxdadG3Do4A6kew6js+cAlM5Tl2a89vqL8KM+7Nq3GbtbtvIDQS/9Wzd27HoNrYe3EbSd2LPvdewj7+7evdixZwO6+Yk6MD6jdB+O5oSiv0VEwBHwlJwuhnCBuMQJBnfmigCeccY4RrdlDPnd2EckPA+W0uho74Bj2VTKQAxSRiKZ8CAjZ2mF7u4uWkIGJakkjzn6mSYoz0dluXzMK8TPrqXR2XaYB2sFUCWkeD7kcRqXl5Qi09uLEjeJTH9fnK+5GHgO/R4tciSPXOWERDqiOLIuO5TPZTHQwQCih6JSVINdCZEgT5sBZpKHbhWlJQw8EQORIL+k46CaOrmcepY0ktNH32cFQmEQ90100xZfjiXFwDnyQ7TsbsGRrh6oUCHHgypXeagsrYyfaypq4GcDjGoYhZF1I6G4a8/0paG5HaivqUPKTSGXzvH4ykKmpx+20ejv7kVlqhye7SLMB5B6jrLhOEl0dfQgaSe4lXVQmaggsFkMq6nHyMYxAAcD3KGXELR8XwYR5Ub8UJlOZxCwUxG/vHraRbo3jUImj1wmiwR39jZ5iw8E93+liXK0H+6Eoz0YHotIGrCdRQ272zqofwRl2wAHLkYGYF9AvZlStsiHif8BilZSUV6F6ppajn45z47rMapxHC1Eo3HkOB7OT0UqVYGqymGorZUvGeBzDb909sB12Ul2pKenDxUVVTj77OmwLRd1wxp4Zl0B+dLAgSffsjivuqoWTqIEDcNHsbyaHUjCs5JIpipRWlYFj4AmaFWpkjIoHus2DBsed37c2AkYPWYc3GQKHj8QZrI56jgGZ501DUpZ1HM0RgxvpGwPMmOS1HfYsJGoKK+FpLaTQlnFMChuVxobRsDWNgwrChbMxNAlM0mH2mCAIoTMyXP+Vdc1oKyKDJwkaimofuRYRJaH8uo61NSPRCWBSZZVw/JKANtFSWUlGqlwlivDuAlnsl4NEuxUsqwcpeUVqKtvYIfGwmOHBCDXS6Cishq1lGOlylBDgEoralFVOxwjRo+nshp1HIyIVuUQoPKaGspKoLysArbjIkW+qapKGE77kWPHIeSgpiinsnYYyypQWlmF6rp6RMwPOYnL2ZfG0WeATDCM+p8xYTK0djFq7HjqUI8QEfunEVhA3gTIhwUEdPYB+6P7CxkMUI5pHqFlwVcaPtNAOwjJqAALyi5BbyFAmid82VAYaWSNQh/nrXESyPHZSpYgTQR6/ALyVKyP9UNtI820v+DT7Wk6TkDR0npzOSrloC+IYOxkXGaUDW17VNRBXtnooepZ6pEXHozJjO3QgUcItBW3ybMspIWCgGcYLuShkCG/NClLPTOkgOWiY4GWFVGu6HQs+TAILIUsQcpog5wFBPRRgQIHX0PPf3wJ3qCleGj5IqzetB6g9axpeh3/8uAD+Oe778Lip1ZBGDe3tMT/I4G5SxfjkWVLseCxFVj85Eos/fEqrG/egrvv/y5W/fRpNO1uxg8WPYrn/vB7WHTCMpX++MrLeGjhfGzZJauSh5/86hdY/OMf4cEVi7Ft3z5kCWSaQP7wyaew4ImVWPHM03j08RX49YvPozef42AJMDbWNG3A1+79Nr76vW9h8VMrcZALg+Ygrd28ke1WYNETAzote/oJ7D50GCpRgs1798Z6z39sJYZowWM/RFemD2l+QPi33z6LO++/B7ff800889wv0M2PBT4x0vOWPopHl8zH/MULMX/RQjw0bx627dyJVzatw61fuQ2rfvwUnv3db3HPA/8L9z7yIPYfPoTFK1Zg6arHsHjVSixYvhQLly7F08/8FK0dHXjypz/Ff/z+eexhPLSMyqx6+in0+XlaWIif/fpZzF+6JK7Xy/hGyuYR6PsemouX1q6FZtCW56q14onHCeQC3HP/fXhowaP43G1fxv3zHkKalrluaxP++etfw49/9q/411/+EtL2Ow/ci9beTry0/lXMnf9o3PZh9uUR0j/e9kXsPLAfW3bvpK7LMG/JIjy6eIDmM2053IqfPfcrfOXOr+OZX/4cv/jPX+Nr37wL/8bB49oDrUoN3JTGdxxiR0gAAAoFSURBVO64Ew9/+z4sue9hXHbeDLz88gtI53vw4Q/PwfepwLAxDVi79mU0VlXh+9+4A5/57C0oeMDIxpH47tfvxL2334FUBHglCfRHeYwbNw6VqRLGP4cBy1BQiLaD+1CadHHWlInk3Ye9h/dBORH9WiX2t+5hvQAlpQkEJk8D1vifX/0nfO1LX0R5ysVvX/wdNrfuxBPP/hRtve346/9+E5Y+/BAaGmqwdt3L2LF7G1JJjz7Iw+WzL8E37voaxk0YjdaDe7Bp63okSmwoz2DS1Am47/v/gge++23M/d53MYy+6me/+Dk0p+dtn/sC7vn6XRiWKsezP/k3JJinHc9BeWUZLjj3fFxx6WWYfekVPGSvR9uhVhj6k+lnTcH0KdNw5WWXAxxdzfn90Ws/iGkTJ/MLaS/kf/s3+4orMGXSJFhas0rIqDZENVexkmSSq1s3jvT2IMNPzV1dXaiic/UYN3V0diBPixg3dgwCWpN8Wg64QBSiAmzHgs8pdekFF+MTH/0YJp15JqR+Op9Fe1c7PAIx5azJmDH9fJzJMvkslC1k6U0MI+UCRo8ejWtnvw8zZ86kPgHS6X5wGQRdKxJsO/uyKzHrootw9eVXoqKsHD1HeiDbqXOmnI0ZF8zkKjgy7ks+n4fO0vwDemwoFVeytEaJl0LDsHpk+9PYunETktrGl//hVjy5YhWmjJ/I2ChkHoEVp8xvVo52IAJMZBgOhFDCg9+zhtXVcauSRmd3J3p6+9DLT86No0bBoePctX9PDOTll1yKkpIS7D9wAAEHRNMBS3uX0y1gcKhBvRw3riv8Q8Y+RAJJLu+WsqAtDUMHLfJZNQYoQZ+n1eBgsb5Npyt1hAI6+5D9lfYF+h75A/cazgpwG7S1eSsqyyrxIKf8osWLuXBEjIc0kOMeZOvO7diw8XVs274NUnn2ZVdgzMhGPPerZ/HvP/85kraD6rIKxikOZ4wCkQCdPkpSKYb3PkS4bVsDqWUxSg5RR4DS2SynWTs6e7vQy71V3fAGGALY0nowBnIKLbE0VYKOrk464jRytCJFPhGDVZtOIKLi8n/x1GwjVjD0nV8AUERENrlSlqcVGmVQWVnJvVcfBMTWgwdjfaorqhFxRZO6bIKNmzZi9ZrVcT2Xg3/dte+HzXTeI49g/cb1PBsv53YpgUgp9tHS6Ge4fvvXv4q//fTf4aEH5wK0hCmTJuPGj30cImguEV3JVSXgKgNeAYMqSoYoKUKVoqpaQZ49z2MR/ZrrYRy3KwluOTZt3oy9e/ZDsZOTJk8mB4Vde3Zz0xpgTGMjZARz7OCWbduhLRshQVEU7BJoRzmsDwhfsHcSm/iBH/MyrBMGIeSSyDmkC5Do+Te/+Q1u/NSNeP73v8ell16CC867ABFDBk35zc3NuPmWm/G5z38O//7Lf4ehNc1533X4yJ99CD3dR/DVO76GF195CYb9IabQOQrTro1rrr0W199wA2ZfNZugRDA0xb++6S/x95+5Bdp18OCjD+M//vBbQEaXjY2lUeBoG9HuGCpwmVZKwea/8eMnoMBgq7unB+1cZXI06Zr6OiprYoBKy8tQW1WNM8aewXp+PFBxdxXEWKABQmDAGQSlVEzMOuEtfkMqdHd302pCtLe3o198EAba1zDo/OQnP4kbP3EjJlA/cAomOYU//5l/xJw5c7jr78I3vnkndu7ejVBF0NpmV0ifueWz+PIXv4ybPnETHMtBikGVUgrXf/xG/Le/+RSDR2Dh8iXo4vGCIqC+MoxL8CeX4wzw8+FjZAOdHd9379uLbc3boDiCo+iDZKp1MnaprKqCw71SXW1tzKdp+2YIQIZvZB+DRHgIkIFSCmKxeIfLZl+uvvpqLF20FBfRER+gb1u7bi1c5UKsSwC67fO34bYv3YaLL7wo7qfDNmWM2L/yz7fj8ssvp8vJ4wfzH4LjJKA9OjRFs5b5GzDMLvDMJKSaC5YsxC2f/Qe0tLfiug/8GZfM8WjhnN7X2gJlucjTeiwyhnqzxuIEIzrbgEcIssIl6aP2tezHjl274XJVa6gfgbbOdo5qBl0c5Y/f9HE888wzsOjjdrGOcCMcTFTMOoptCHxWsKCZgiQTUKwLg9dQC7AMcG0bo7l9mTJlCnJcidoZn+UZeiilYGkLEadwjqukZVlcRPoh7uX2r98et/30p2+Gof47djSjwD7oRJiEpxKUZsDhRMH2kUEBTTxvWb31deztPIREwkWFdlFme2QO5NnQo7puxgcoNGQntLYgfw1m6BMco2EVDBrpkB0vga7efuxh2FA5YjiQTGD/wUOcwgZlbinGjxqH0Q2jgFyIAzxBED+X4CoaEowebkI7eJAWsMO5bCFeHNzQgm15aOvrRUeYg7Y1MlxJS0uTcKBg8gWUcspk+7tRyZXOuBZkMBNs41k2hAK6FZehRsj6GU74jVygXmKQmeGqKbJtx2GYUYCnHeiAneztPIJvfusufOG2L+Dzt92K1Qy8JkyYQEVszGNkfcdd38COrdvgcpRHsJOGgDjKQlmqFKCD0ErDJkAWyePyDI6A41jwaGnjR49FRL/Uf6QXE7ihTdDUWzjl0t09uPWWz2De3Edw//fvwzBOswzDiogKH6FlyWr1yKOP4AtfvBVN6zfgjMbROGNEIy469wL46SyefPwJfOc738arr6xFA8+JykvLoekbE5SvONaiWzaTgejGMUSBMRe4+GxnP778T1/C57/4Bdz82b9Hd083xo0/I7bmu4jB3ffcTdgUJp05CYYOXFt01RL3bFi3AU0bm9D0+ibs3bUHH75mDq648BLs2NaMNatfRnVlNf7py7ehpqwajrKR53FnRGspLy1DSEdsaLZimiGdu5BNACOudpNGn4GI50k1iVKMqKyFYbv+wx1o4HFGA5dfh2ZeyfBhGI9X+nhYdmB/C2qra+h/DLZu3oLmps248Jzz8Hf//VMYVlKJ6664GtfP+RAOUsf/eu63SLgJ3Px3N2Ns41j4OZ+dQhxiJLTHqL0EtAEUeDY1dChnaY1XXnkF69evj8nP+rjpYzfi3OnnYO2aNdi+dSuGE/Cb//bTnCMa+qmlT+DJ5U/g8ZWPY8XSlVi4YDHmvG8OJBL+3p33YBkDpvlz52Hl8hW4+orZCDm9aDY4Z+rZ+OGS5bj1Hz+HEjtJpXycP/1cPPLgw3SAX2IHOWMJ0i1/+Tf415VPYtXi5bjlU/8DpbSqT370hvj93MlTkdBuPHW+++17sHj+QnZ0NO6+45tYOH8Bnlj+GH7+9E/x8AMP4hMfvgEBTwBG1Q3Hbbd+EYvnLcCih+bhh0tX4KY//yQHKcSHPvBhLF2wFH91419BfOFH5nwEjy9biY/9+V/gHMpatmgJFlHGUz96CqseewyrVq2i9dVh5gUzcO8992Iu5Tx8/0N4aO5cTJ96Diyl8L8BAAD//2MSS9cAAAAGSURBVAMAqiFtgefXGIQAAAAASUVORK5CYII="""
LOGO2_B64 = """iVBORw0KGgoAAAANSUhEUgAAAKsAAAB2CAYAAAC3SGPoAAAAAXNSR0IArs4c6QAAAAlwSFlzAAAOxAAADsQBlSsOGwAAABl0RVh0U29mdHdhcmUATWljcm9zb2Z0IE9mZmljZX/tNXEAACciSURBVHhe7Z0LkFbVtefXOd/5Xv2gu4VGmpfaKAqCvBS1fUaMMTomNw+MmSQ3U7eSTFJzS6syqUmqZibHj6pbo7duzVwzj1vXWze51zzGyNWb0ZgQRA2IrSK0IkgjSoMI3Q0IX0M/vtf5zpnf2l83NNA00H6Ndvps6qO7v3PO3vus/d9rr7X22ms5qVRKwhJSYCxQwBkLnQz7GFJAKRCCNcTBmKFACNYxM1RhR0OwhhgYMxQIwTpmhirsaAjWEANjhgIhWMfMUIUdDcEaYmDMUCAE65gZqrCjIVhDDIwZCoRgHTNDFXY0BGuIgTFDgRCsY2aowo6GYA0xMGYoEIJ1zAxV2NEQrCEGxgwFQrCOmaEKOxqCNcTAmKFACNYyD5X74x9bVGmlVqzwy1z1uK8uBOtHgIC7r2+C1FdPEs+JmGq8TIWkflIrsxstd+kthyRTyPKlJVGnIIn0wVRza89HaG7cPxqC9Rwh4B6qjUt190Xi25fKxJo5UoxcIZbERYJAJJbg5wRxbJFi0CURyYsV5Y+gR7I1re7Cm3eILe9J29T3U12PF86x6XF/ewjWs4QAII1KNLNI4r03w0mvEduazWpfDxAvkICzbJYV8LEAboSPAMqiBJYvfqBgBZj2IbjsIfGt7XLx3tWu3bQm1dL8wVk2H94GBUKwngEG7pw5trzddplU9i4TP/JnYkUWA8o6cWCf8FIJBommylxtJIIIH72upSQgRPm+AhDPkEJ+gfj+NYD4FnfhdaukGHtVvIMfpFpbiyEih6dACNZh6OP6/kR5Y/f1cMjPg7rPiBOZAVAFoMEkPYCqD5v/SqUI3nw+RQCs1/UzUCzFIs9GIvBeaxbXGpEcbhTH+4NEJv6Lu3Rpc2rDhlwI2NNTIATraWjjBvHJkpUvgb6vi20vAGiVRiz1AaL+PNdinlGQ87wC3rYBbWSW5Ap/TqXTxbMnu3ffvTb17LOd51r1eLk/BOsQI+32AVRb7gNU3xLLno98WgKpctRyFAWucmH96USq4MafAcEXyYGuue7im3+Valn3Tjma+VOrIwTrSSPq9s2ZILE9y+GA34P7XXFseR+NkdcJoNw2YsclElsg+dwU8QuOu7jp71C+9o5Gk2O5zhCsg0YPjlojzu4vSNH6DgrRFUZ5MoAazdLPZVU0iMUulFxuOW3udKc0/SrV2YydNiwDFAjB2k8JNwhq+fVL4gV/Cae76ph8OhKsRDAEKPjOpagy5kT1uUYe+7JMsXe5dXPWYiUY7dlyLr38WO8NwQr53dramLR33wJSvoXSs9AArThIkx9uiIzSpTcck2l9gM6uVczCfFWSTRWIZ8Ohtc0Idi/fvwXDQVqSdR1UvP1jRcgnqPEQrDoYew5fLNGKuwDIQuRUwHUWm0sKUv0osOOQMeIEbKtaUpE4giK2keW8DoQig0otO1pVxvZqTFvDMMoBxSser5BC4U62FXa4c5BfW5sPfIIw87F1ZdyD1Y3HE2yNLGQErkHzTxhAnamoUUBX+WhMpLpCpKZaJJGwpLpKuehhgPk/xOanLyzpiBQF73bun2s2BnTTYECxOl07yolt+wLstfdKUlqxwT6FDfYsWf2ZOj92r497sEome5HYVTdLULzUIDAYBqwD9lXlvpWVIpNgnnUT4KyAVkE4AdD2dOfhsu+kNqxtwwNro6x5+Q9o+aupGE7p3c2EuMIs9cOJGQpmFSEs6xK49PVs774MxPaNXZiVp+fjGqxu7VJH2rcAnuIiyFl9VsZ+3UWtrRG5cBLghJOqw9Vg86uPvBoJDF1xE1TkH+bzorts8ZuSjr3Nzd8FqEtLgD3NxBjYfLAkAXdeIAW5gm3fjvGubI1rsEp6Mx5SMhtT1Uwjqw4rT+rSbwdSVyvSMBmgVuK0ctIzA6D1jJBwQkk935J2m5qeYldMBd0IEwNnGHYbTtembkCY68WZEocbz2zcKK2tR8rDo8ZmLeMbrIFzIcb/eUBr4rDDZ5Qp7qhM6tIfIKOWwHg2Gv6gilPNzT1sqf5G9qVrWN4vlGj0olOcYQbuNyIHzVgymXvnyMFuWLmEYB2b86wMvfY82KSFc0pE/VCHrzDKcq8grawYEVAHKmfvv9udv/R3KGeLEQO+Afd0jOPLKaXflyBAFJCgHsUNAXl8l3HLWV3XteQ/PoRZSWrwA7BOi9UBpSqBf7XKqKpMnY6jGhifxW6A171b4hPXAtJlbEDMNI8M5RyjeLXpYSDTpGBPd++7b3Pq8fHrtD1uwSrfeagS97zLxY42GH41nCeVsaUC1iQf49QyBBfWeyKQ00aQDTjKMkxRRcldvPQNThG0iO9Nx5KAq/ZplC11KfSKDXjETsFCwExB3RqnZfyCtSFbKfmKRoCCh1W/V9XpmGKMbdBqTFUJxYqRI08sJY4qku1TIO3mkzkjnm6YvVNe3fM2XPpu46g9VNtmAiB+eEWMuX5C2tpKrtzjtIxfsOoC6xXjLLJsbzL6w5mRcIoSBSyPHHPtOwYY9U1VsBfZBCiukUj053LZbR3S0izuK6/EkTVP5bJr1viyb0pE7mjsk5jdw24VvrJDyhZqDUBE4QxXYBWkDuVuHJfxC9Y2ZX5p35ifjPf/MDbPShhbFFIV+jeRBkPGcD8D4gMctfqdfPebv9Nj2O7q168X27lafAcFSU7cfbr2zkDeWW1J/bWOLJn/M+RglvohUGj7FnItldsfMk82S3V1fhxjdRyfwWp7G0eTmG3sqwasp9mzV1m2ArBqyXDqRLnoyfKtAr7o14jlXyv/6a+63FffRMMP/i0ovoHKYcknbYtF4nD1fFZef/2fpDr+t3LP7b2c8zp1ia9LBJLOWlKXKEq6LiOzZ4dgHZezNcnin8FjJeCfnkk1YB1CLzI+rQD2CCbOKqwB6g8w+JDgceJhN3WWS6F4kzmKnQ9mMREwzFLY1BosNZQ2E5BruzN5+dEDnalc6uyUphXPjsuhGnjp8SsGpLO9UlW3S4oF9Wi6UDhJPXQBaLqP34fYUICxqVUAMfLEYryv4NKRSQBxkvETMH4GQ9Spc4KNMO7fLU6c7Vd8q8JyVhQYv2Dd9FyP3HbfNjicHh+50IgDw5mvcgA1ixigDiynMOBBXyjXNfb8fhfCk4dhwKYayCZJVrxOVvLQufqsoDqO4wakUisC94V8F3bOo7pVf9oyAK4MnHVAFFDuOqQRvx+gpwW9mr0U2MFuOPBLsucgJ1lx9grLWVFg/HJWJY9XSJvlWIJec9R6uKI+pofSKFvcNplALGfixEPVZZ7x+/jvWbjv86mvXT+uFaazQuigm8Y3WCWJuSn3HkDtQcsvxQUYqgwY7PvA2cGD7GShN6mb4LEl/QzmzwHzWOBrZIxVEvEfk7VPt8kdC891vMb1/eMbrFmMrRXT9iK3HgSoF5aE0WGAp2BOYxWItJfuq8LZevDhQAXvyZ9jAm6xlzbWYnv9n7L26xtTqWvGtYF/JLNuXIM19bnPZdwXXt9IDKuXAdJFAA0H7GHIaHxeUd4//BDTFMpWPX6ttZwUiGHOUs/+kiuh/qcRLPRv1baQiWUn5q8N/P4vUplpTqXCE6shWEdCgZ7sboz+zYBJDfj4tp7G5DRQ94CTdhccNo95tJufKscm2KiKoXj19cJBZT0A5YIB707+W8/jb8nLz7SrYjeSbobPjGNrwLHB3/RcRm754utoW6/CNlHN8W09UyyrAeUqi4VAOWwEXEbZqFKwFvIfiuX9d4lU7mDLq0pyckg2/fkBw03vWBFi7iNQYFyLAUo3Y8La6uyQdOIZKWSvgRMuOCN31QcHlC4TMRDA5lHs9VMoZNhqfS9118L3j43L51o/whCFjw5QYNyD1QB23ryiu3XfOjmQ/Rky6QPYQC8xBDoThx0A7QBwS9EBAXJRfQnDUmYKhGDtJ2hq3rQud3X65wRGY9tUvovMOfusOOzJA6LRWMIyKhQIwTqIrKk75h12n976mMRyPoD9HoC9fESAHZWhCisNwXoSBlKfm3fIXb31VyhbeGJ5GqRtVkmrD7fwP+7pEoJ1iBGAwx50n37p/2LSiiO3fhPgXoYcy8G9/hOnH/eojdP2Q7CeZuBTn7tpv/sSIkHG28tW7D34u96I5jSt5BMwoHyFJtPzOW9CsA5D7dRN8zrcX76yUhqcbaj5WzBJ3cXW7OU8Uo0si8Z/0vas+gCUbFqhkjUKKA7Begai9ntGvYFYsAexYJOJCij+FXDa+YbTWgSg0M2VASZrdsBGYaTCKsfxGaxzHHzEApKuySpOrL4g3eSzkhgcNlgIYOcD3IQRD4qFiaQN2iXRJO5ZYSk3BULOeo4UTV1vfFB3EtVvlzzy6KuSSUwmwA+B1vQQqgYOrsINMKvADkuZKRCCdYQENeEn77hJw1nq51jRrNgcxQ4FgRHSdbjHQrCWmaghUMtM0EHVhWAdPdqGNZeZAiFYy0zQsLrRo0AI1tGjbVhzmSkQgrXMBA2rGz0KhGAdPdqGNZeZAiFYy0zQsLrRo0AI1tGjbVhzmSkQgrXMBA2rGz0KhGAdPdqGNZeZAqMGVt12lN/+cQ795eNp5ogA97pAbL9TMum32K48tk3JPrtNBmi9F28mIkWTQsIEiy5IO7FS3pIldZ50dF+Jt9PF1FOUgrUltWWtOTJKlsC4zHDIEognlM3BZ7E3pVqaPzDXmpqmSJ48Vx75rkSOSq5vc6q1ZY+5tvhm2vKu5ldfis77hFN/m4S+XeaaZnJ57rl6HFbmc1JgCk4qBLnytlKveXagmLYbnbm0eQV9tSVqe7yfxnpVX8EuUm1SZ8vxU66Dn21qapA+WWT8CQxt8IRxiDXge1toZ8hnTmh76VKNrDFX8valtBnjeYt+Fkjq0UZft5DrNecuvuUy4sVeSb8q6JN/zG1J2/K9I1Ihb5Gb6wPoH5Hq+vnQgPyyti85e7dUQ4/m5u4Sre4mymLfQmKD1UqF/R5Dk+bdLiHhh0ZfZEz7/cx8jeVpimb2znPtXWnLbk91bWBcPnoZFbC6TctmyG/X3MTJkC/ginSdSeBrEuuZtHnvSrLm1+7ixU+lWlr2ki4nKjv2NPEq3+aWmzhzXwrAqx6hEXmflGXPysEjBKGQa3mevFGQOBb8wl3W9Nep55sP42s6RxzrAWKifoZAFYd47q9Jyf643H+/J6te0KPV9zOYCyAeISar/hu17gHEEyTr3SfRyPfwUSWYsLeWbCiPcK3ZtP3EcptMgJcxuHoO62bwjD+r8zf67Akkb3Cq6Nfnaf9bEg+ixHEtUp/D+2oegA8lnljnLmx6WirqXko1P3vUDPycrQDjgUslk/kGf36JQMPEdIUuVsAA80zg/9FduPQpmVa/jpxZQybSIFN2FYBZxrv9O569BmLFOVELZIIMk2a7BM6TJCdeT86E68UJvs0YkL6IJzzasNTjhslkwQb6gjUA+jEidewhed3dJNv4S/pfJEnyU9JN7AMumPd1emZL3vuBOZPWV/gF77sFIN5L++S85bwPb90fjqZokjRZdhSa877E9Gpw9jJtP5lgNbMwm/4aXf4WEaA5vwTqfJ8I08xs24lDkHop+JPFiiUB7D9LezvhTDg6YkW+DAHizPh9EAAXO3syzxOTH07l+QSM8In4R1bpyqpp0n10GWR82oArbs8jWs9tUlnTIL09U6SYv0Ua566WVavS4iU0nuRVBACul0xfF30xHJeykM8dkqikH8yf7qM3ce0VQKycpge0wuleSAKEqfiw1ktv98VMHMJen1yycKhEPdGwSycIenmU6AGlialBhYPL6P8C8dL/h8yC/w/wHRHnL+ZLoY/Ts/ZX6Fet5LJ6ew76RAlBdAET/FKwdLl0HNQsA8+f3CL1JKUjzcRkgkbsm0zY+AIRvC1exLYuMH3JZSfDcXnUn0Sd0yQRnyxZ2vGYTBbcNwCwkchEfl5MGiTiH1X8kvYnSawS+oG7XI6shknN+NFfggqGcQZJQGYQhYYEckESV0iNXc+zGjAp0IBzAJSjPw6PaUjQ3p5axrJG6vJlc0QvK2eFo0WlcGQRiXb/DfGfZrFsEOfJb+Z1mO1WBmKp0/KnIOilEPhegp/sYZkmKBpcLE44k1yW3/0nJe5skWyxkRG/EkJsYRnfKFF/OsRnCSpM47tp4lmXu4ublNM1MgBVJiq15ozyI/PgLhfhzE8i38wlcOIayWUOE616ferNtbtY8qrgmp8ydff1Qmgi+1kWQauCmyUb/T31vSOykrlSoxzDI2hFCYC+d+qJQYQDmUTwKx3gQqGbd9wA+LdSH5yE+n3qjDjXshQXZH/6EP3dwvt/nqe+TASXWgC0i3ZfZhJydMavhT5Xs0Is4doNtPgV+roFcUkjcx8vB3sWkr/gL8w9Pn33ZB11vIGLoibLuJSoiOShtVqgEQHnfAUTbwGWit4u+vUKoNtNv2AK/qeJhniV9Pbexnfv80kARAQ1zQ5DrdFBQRPyMBqLHmkE8MBkQdZJ/yT9fgv66dKvST6uos4l/F0lGaUrdIhyuiKtgW3LU8oKVmlsrGdQWa7t2aZ7QfAaYPlfYuU3mpeVBN8HXQzuvRBuLgC6nhd8VReO0mE8zoVouPO83wvBN7HEviHF6Fbk0x2IFhm443YG+EYAMQHOMocBZtkMkFUdwlXq42YSk+DMXyiF3io48yVMGiZBDg//4mbTp2T1Zdx7NfdOgKsQv9LabyaLWIsZ1AUA5N1SUrXrtD4lulbcH856CKJHBuK7K+eX1VJhrZT04aOSnLiEd0WSte42WVty+esZPJWQr0M8mMjk6qD+n4vl/JqArx0ibfQncQ9LdyXX50Cj2yVe93uW82eQP00uFyPb+wUyals3AFYk9OwGJsP/lt7MSwKvI6zhLOg/lfaYcD3tUoRhwOtKSeaC7bT3KOLKRsnCgS1k8QCAWSqGBDNLAeW4UbN6D51v0dxhkh9nPJiE85rUsRqk0xGpmbhY8sXp1KXJPjTu7Tbe6+dSHbwody7vkxXlCZtUXrAmqqYgC85nxk0EhEVm+1uyN3g51dnS1T/MryHDkdS3uJDlaREDopzvTQi0m9A7i/iOQfS+DkBu46WZ7cFbREhJs/S9T1zUDgliWyAoaXacyczyRTwHuIPLDaC8PMqJLksQX2VfR8jcZ11kuI9Ym+HOiBJaEBssa3Z/3qk3zCCSZpCflyG33g5P0aX3kMnjd6YUaXVmRurgahYXfUJ57X7A3gewEFHqLkEGvIrlmuRw9MvTwBn2NJOILeu/xURcmdqyDnnYlLS7eNmLUuy9FS7NfawOVgB4q1/jmsbYZLGoZiLLJOpjtcj1wjOfl3TPi6ldLdquoNTul5/scFJdjxcQaTT9u5lp5tnAUlma9dmvhCZMcGij4TpVEQykE5rUmNk+VNRPG85qOCpwCRDLnESOfpdk8KamSZIpNjFprmNViDOOOxmfn9LOr9EpOuX5khpQjlJesHoI+j6yXQwq+LpmFDPSaOchxfFi46zsC5xTj4F4mv5kPwL9YwAvDVBRsMyAzuX5K6AMHMS/jozQfwfXfkp2tL8JsXQ5Uy10HsvedAaeJR9h3iu+CLF5xobjIssFAZq+dTHckvrtN2T5bR+6a+omSfbIEtpQTqICZgsD1sI9jchZl/NzmVROXMT3axBVCnCxQf0+puke/y5J6p+BJMG6OpTywJdounx5Rp5+4QDv0mMywfhA39KDhgoYA54eiXonOG7zbTcgUqUGJceOsfLUAVxdYo+XQNNt6jkvImhbdruknZISROn3pR3odKSUk7Z/Obft+fz+A3LEdPLuiAvBAslmeqHlS3S5GQXtFjg2YpT2DYtLFvn2eGF26arX/0V/Ijr3vtq47PC/yCr3VVaNBjjqu7T5KOnFfpna1by/HAAdXEd5wSrBEV6I5YdIZXr6UwP0pkX5jzmTxJIWg7ss4FdSl+uqw1XLI2vJTKJPt7+Pxvl7BmExM5Rz+hZmHQBp2zfz/SGAuoXlZzvSEfKZv9QsYza2gig5pXK5t+Fg/wrwruYDtwbAOldKYdHfQ6FrNYnUFl4HgO0mltkEilgREMGVPTgOMqtZ5Cw4YfErRraEVRsxwAyQdlbl0JNKoi4vmcMlc42logKyqeehyVDWbKrlBTlUGJlCnwL6kAOI+8ykNMspsnbMXsxvJa6pJdNdI1Grlu85gFjsQbHcQH0Egx1UbNMXjQeLjFtYJA3eNHijMXW5P8bktvK5SsxOWZk505P39ipvLfVPglo49bUAPY4IEec9kXf9NSzXPxO78BZ16lFzOKuKPsWoVDv9z5mHp5ae10liwiMV3Dk/jsj2VXfz3bdJqXQpY7SX9/+pTLP/AdPbkVNoVYYvygtW52C75CfC/YqfllhyOgC6VuLyZ+7SW1QJyCFPLWSwvkFoyBlwLV1Gdpjl1m9fBnedBcAgmvxK7Fgtf3+V36dwL9aDwlSAOhE76DZkyTcZqaNGQdEhz+bggMi9HjKUrlOOdYiBhIPxl+dproDNEk20sVxVS28A5w7mmli/vp9k4D7NUN5qTGtF0laqJpsr3ojN9jrJ+UcAvdcvs2JOc2a5N988F3kP9QPxw6Hn3kGCYLCmKyB8ppHvX4IFYoG7tKkHDk5b1hcZyMlSyB1E8XsP3voOFoIPWCqvhXNfKbnCv+d9GIPYdvqdRMG5i2euQuvmvYrv8nq/553NcmvKnXcWMQnupU8HoAsau38nKeV30x6Wkagnv3txjsTyyN72LmlrR05HpkTbNe9Q9LczaQjtiawfSJOZyBJ0kq9rB9aKDknW76X9LkPXIkE9PO9Wd37Ta/S/Tvz8p1jmkXENl+6TpIco8cc7JRL7vjiRq40CJ/77UrT3Soczhck+Her4iFWHpc07hJ11qPyJ5wzf8oI1vfyoOGta6MVmBqSBZeUqZvB/ZQajyKiCJfXIMzP4Tk3oL0g8shrCwmW97wMcFAfnA15xPcrIbpaVy1CgJpj4p0L6nyT8Q0sBWc8pbsWSgKKlBM+yrGExqP70B9K96nVAxDU2DxR4+TzLrLVVrqg/KG175lH/Usw4lVgB0LDRmG0sFLqsAlMmESYjOK1t0z9ZyiCh3Hn9XNOejpjxfelGCw9UoSK9S4GI1hJ9GRQQqx1bo2VPBQzfQqf6ApsWyOv+hQzkhTDobup9BrA/Jw3Vu2R/N3Kph8afANhyF2DiaDfiQAHlxFLlMHoBBndkUP9fERr65ezSuOoyDxDWMNmuYgn/Gu84Hfr8AJrdp9mIuWUiE7Wett9ngv/acMFoNDCpPPPZfVz/Bd8hAmE6TCZnU8etYLxFqut+CgX+KEe9VwDcZ3mPxYhhK5hcrAKFGJNWxS0sPYX3uM6qQ9p58b5DhsYbjPxtRKEA+gb/mb7pgCWw7zLZg8exs/4zI3eyuHPOQNUHygrWVCtLLeYWidY8wYDoSy4y8qUOmhYd+qKRL18FyBijM2+IXdVkDNEBfYmyI1MoqDKCyQsxwkebDjwUEOsP7HrtMnVU17/N0ruO2XwtoNe9E0xbsa2pDSs8uGenZINt1HEL2nYlg/8uy9wO0kj6bDwsYJAuRWSAW1JfxPlbSSDP6qKdRzFzgq9S53+gXZ4jO6AUO43SVCgo8XkXiwEDHFqMggRHV5NPYPXxToA+UDssadajU8xyWQDoHiKOFTwPF/0HuedT28h5FcCtXmRQpwOULwDwS6ENA48Mrc+o6cgr7KGuVfT7yVTz+lOOdLPEvkf6959yHysDJsIIieJsu9b0S0FjkhKrcmrt4YMtFcE7l+eCrSakQ2Jn3xM/vogJ8U3oezHf3QENXoMW22AOv0WSwZqAndeOzGCMOHJOv1SkKvAuvo/lwlnNjl8jDGGimR8eNl5fzQ1BNX2pMRxbszBm+nIwqfXlxFhZwWpmf2vrEUDzW3ZAkGFYblTz9tQOp4RUuQ4ZLRp5QWL+ulRzy1E4xUZm4yO85CZeXOP6sxujuyzwN8uoZhvgqi+k3mw1Fne5/eoeWfX8erjmJQb+qiDURZF5KZ7HDk5sDVuOqtlOoB6I1fOuMPQoEQDKw+bLPdHIb6T34JuplmOx/fe585c+b6wRxWAilipVFFCOgpfoBZGszaQGfPoOuknhKTfuAxQoaJECk0aVJ0QCS+VA5D5jpcDuijXDiiECZQ1QTR8/8N6VGYl/lEheJ9+1ABQLBhxQsxJaVhcA2kzba+Wxg+/IPPPEqaVtx5sye+7fo4Dtoz3SbiJ36wrhA7uItCGavCBWHwnpEguhLXqAvpOlYlIfu4bt0PyfeAcUNGzBYu1lya6GZrpr8AwrWpZV42roVMf1kgyr0zmwtklN7Depdet2s1WtVoXViHLYclHGjKnLbP6UFLRihlUCRdGKbERjKcnwZShlB6v2iV2gw5hRXpSVa7YiqE+CCEw1nYWOLlU90lvTmdpY2kqEU+xnM2G1NDSyvFfVAsxBJdsn3c4B7ikBdaB4/utw4dJulNeTlpnb0rpRij0yz0Rppjnd649LxkI5yRyQ5ZBwpQdwvXekOtEnTvX+1Mb1x4z8xn7pdW9BbutkkiUk6UNg1bKjG6SS3Ru2DdCi+wvzSPvoUafVwz0JH777srnHQz7UYq6jkFUUDqaa13UN7rqR37qIOzCl6UOZymSIWgn6SUzYAmtLsle6DxxgwnedFqhKs64uz22cuUn2dO9ltxAFtuhAWzrm5CSCSJG7rdOsck3LXpdutlJ5Eeh1BJah4pjIPZ/eKk88cxh6T5JoNmCV+VDlYcSMD9ymOU9Kon6d9LDtbfwdEHEc5OFu5Nlbb+uUdcw9jf/leL+gTmTXExSx0qtGNW29jjWK8e3p3v5N7BOGcCR/jApYtSP9ZhTdfTlxB2aIXirxpatFObF+hi2pFcqhmlVDPq4lq5TcX5goKFUnynr9RmmV2fRzSjExAMz5/+PONWfqx0nXh6x3uDpSnWjMnZokY2Ql9fjjqpqehmYl22aq+XmVFU+RFw2Xb205kR4tA8+00if9DFFW9N/TamjcdlY9HzQ2Z3X/MDeNGlg/asfC50MKnEyBMQ9W95Va3mG2pK7f4JnlfP36qNx4Y6GfW5bc/e5/xJGOBh9xwOzocM0YvPXagCzpurg0Ll9py8rlvkmKoX8/sioiDT8NUq3zSvfX0lbD+oD7fPnsnXhPVQeaj+Dkuszfpi8dUZmasSS5wBvKfGNc89Z32HJjg699KvXhXv6+MSJvZ2xJNxZS8/he66rucOT3OwOZNcvYnFiN8q7L98tXWsf6sJX6bqy25IHPFo/JyNqXP5EoMWMarG7tobgsiKOh73DwLcW2uBm/2AsukU1tve6tS9vNnvojj0SkPTNbZu8IZJvXjmue7y5p65V7n/DlkVrHLQ28Jz98GF8CzD7LH9adl17z9wEUvvSSo+6DP9qLTOfIglsvl/Yre2TNkoOyxJuidlwAr6Yzkfvuc9zNm0ugq2VnZ1PbTIzzyLoodpnX+txtb3cNgOoY8LfhD1Anlaa/D/84TZt4Mi2ZLDszRalnd69+h+329e2UtrYA7b1eLpiakcOZJP4RBfehB/fLoQfrJLkk7r6SPJC6nndobKuTA+gI9z94mH4dNNaHpUsdefRRB8DmdWNkLPPrMQ1WaU9iz2MnyGcAM3s6sR1mUZDQVDFav7ZZM/t1yfqGgEGsRgEh6p/TLsnENrhRt+Fgmduj0p4tYp3ENpOulmz1LJnroTRZvZLI4p2VmMuVdvxb98mmJXFMO/g0VOyTBElcPQd/VwdgY1hfudKSurqodHQoly0yOepQQPAllVZ0YTXqV3BdVbTjpqgG+nVgz2RUmNnc1yn3P/AGG0XqlscultfFT5Qu7LuJqfvpwVGpM0o1+n4SU1pW1TmRHyZQXgH77J2H3NpZEclnLkZxn4ESuVOeeOKQ6AqyeXOUPsTlgZ/oe4Zg/dhmayZpSyUaKc6XeP9EU7ncETca3yMF5xLcK2e6ZKRiGe12+4IO8Ssa8OZSm2aP/OSRD+WJ+/Pyo4ctBrJ/xzuJ8zOgUuOPempaCd1br+Y7dPvlInXpQI6k0bpxgatL4N4HSNh2MiKDzDGuLHDtkjUgkY2zkzUZoG/juV5JJ9k4wONxsA9yR4caa7ElczJCjeizZrFRnVHmUc1fCnBf/Fg15tGYzEoG5OnqB1oiS/ue4Zp9OjHxEb5yqj6Ln6m9BIP+HExvM2Rb22HebK+01TGR0jHEkSzND3J2+NhGbcQNj23OmkxwjARTjZqpzJ4fmHF/2CX/5a+OAKapcNoad86PemXTQ3qcBu8qtfp5yIPtcZbEnJueUpQ9WV+mmU0X9u6pK9E/oEESww1mL9vpgmMid86Ga2e3ip+rxGNqonFP9LwOzFZaLJZqXxYsKMoGUrRaOKT4znbgi1mqvgGDe7dcmcxL1yD3gvaMTgma9T4wx0R2Io/KBZZUxgAxW7TGABaFw/b0YTaL0St2p7CNOt0Od/cwofg9gf8A/W9zIkyQCrH5O7DxoMIy352vMK7ROsmMdxi9GeNlbIM1w7J/QWY79j5b6pb0SdcGNZn5bnV8L7ZAll31DXiCIdrGtkSiS6xsBxzIlj1T+2QaOEjXFYQNVMGsye9HJdGxnZ89MpG/sxl8UmUT9WTlgZkoXRt8t7v2LUn2XUCFVTjDHJG5jR+y1Irce6/PspvDplnirHPndiGHNgP0WjW4At6c3I+Ct4J6B0od58rS6W4mVQag9aBU+UyKLmlPs5NUxbEV8mumE/tS06weN01PEmzBZhMFwJ3ENlpw3Qct+eFDiCGANzOV90tyjCbzB8y+OcAc5+s+mYBUMpVJlwDEOLa4d945psNxjmmwpiZ26bJWsuMC1IGCOJB1U67xZjLLZdvb+3HZQ6vHnjoNhqZA1Wto2kiVpd+ndSIUWnz6r1XkVMYkxbteLN1Ee3oCoMPdujVilCUFan8bRj7sL6adCjnszmnUyoyHaGrFSZmwf/SdXmTnjNwI125oVMXM7P3gvH3AnTMTG/Jymu1XiB7+YVaeWDlgyy3VZ3bE8GmlIffhB/S7w+Y7Zd4VcPYS3zYbCPzwpJPNQOX6Y7iMabAOR/fBppvBwCrHWA3W6oftQwmAQ5b+7NhFMze69L/jpQTc49715t7WAVe/U6sb/K7leL9Pah1/smD9pBI87NfIKRCCdeS0C588zxQIwXqeCR42N3IKhGAdOe3CJ88zBUKwnmeCh82NnAIhWEdOu/DJ80yBEKznmeBhcyOnQAjWkdMufPI8UyAE63kmeNjcyCkQgnXktAufPM8UCMF6ngkeNjdyCoRgHTntwifPMwVCsJ5ngofNjZwCIVhHTrvwyfNMgRCs55ngYXMjp0AI1pHTLnzyPFMgBOt5JnjY3MgpEIJ15LQLnzzPFAjBep4JHjY3cgr8f/MhKzEpg4unAAAAAElFTkSuQmCC"""

TEMPLATE_HTML = """\
<html>
<body style="font-family:Calibri,Arial,sans-serif;font-size:14px;color:#222;">
  <table cellpadding="0" cellspacing="0" style="margin-bottom:20px;">
    <tr>
      <td><img src="cid:logo1" width="90"></td>
      <td style="width:20px;"></td>
      <td><img src="cid:logo2" width="140"></td>
    </tr>
  </table>

  <p>Olá, {nome_primeiro}!</p>

  <p>Segue em anexo o comunicado com seus dados de acesso ao <b>Serviço de
  Segurança Humana e Digital (SSHD)</b>, para os sistemas da Prefeitura
  Municipal de Santos.</p>

  <p>Essas informações são pessoais e intransferíveis. Qualquer dúvida,
  fico à disposição.</p>

  <p>Atenciosamente,<br>
  Complexo Hospitalar dos Estivadores</p>
</body>
</html>
"""


def normaliza_nome(nome: str) -> str:
    nome = unidecode(nome or "").upper()
    nome = re.sub(r"[^A-Z ]", " ", nome)
    nome = re.sub(r"\s+", " ", nome).strip()
    return nome


def extrai_nome_do_pdf(caminho_pdf: Path):
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            texto = "\n".join(p.extract_text() or "" for p in pdf.pages)
    except Exception:
        texto = ""
    m = re.search(r"A/C Sr\(a\)\.\s*(.+?),", texto)
    if m:
        return m.group(1).strip()
    base = caminho_pdf.stem
    base = re.sub(r"[-_]y?\d+$", "", base)
    return base.replace("_", " ").strip() or None


def carrega_planilha(caminho_planilha, coluna_nome, coluna_email):
    wb = openpyxl.load_workbook(caminho_planilha, data_only=True)
    ws = wb.worksheets[0]
    cabecalho = {}
    for cell in ws[1]:
        if cell.value:
            cabecalho[str(cell.value).strip()] = cell.column - 1
    if coluna_nome not in cabecalho or coluna_email not in cabecalho:
        raise ValueError(
            f"Não achei as colunas '{coluna_nome}' / '{coluna_email}'. "
            f"Disponíveis: {list(cabecalho.keys())}"
        )
    idx_nome = cabecalho[coluna_nome]
    idx_email = cabecalho[coluna_email]
    mapa = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None or idx_nome >= len(row):
            continue
        nome = row[idx_nome]
        email = row[idx_email] if idx_email < len(row) else None
        if nome and email:
            mapa[normaliza_nome(nome)] = (str(nome).strip(), str(email).strip())
    return mapa


def monta_matches(pasta_pdfs, mapa_planilha):
    resultados = []
    for caminho_pdf in sorted(Path(pasta_pdfs).glob("*.pdf")):
        nome_extraido = extrai_nome_do_pdf(caminho_pdf)
        nome_norm = normaliza_nome(nome_extraido) if nome_extraido else ""
        achado = mapa_planilha.get(nome_norm)
        if not achado:
            candidatos = [v for k, v in mapa_planilha.items() if nome_norm and (nome_norm in k or k in nome_norm)]
            achado = candidatos[0] if len(candidatos) == 1 else None
        resultados.append({"pdf": caminho_pdf, "nome_extraido": nome_extraido, "match": achado})
    return resultados


def carrega_log(caminho_log):
    ja_enviados = set()
    if os.path.exists(caminho_log):
        with open(caminho_log, newline="", encoding="utf-8") as f:
            for linha in csv.DictReader(f):
                if linha.get("status") == "ENVIADO":
                    ja_enviados.add(linha["arquivo_pdf"])
    return ja_enviados


def registra_log(caminho_log, arquivo_pdf, nome, email, status, obs=""):
    novo = not os.path.exists(caminho_log)
    with open(caminho_log, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if novo:
            w.writerow(["data_hora", "arquivo_pdf", "nome", "email", "status", "obs"])
        w.writerow([datetime.now().isoformat(timespec="seconds"), arquivo_pdf, nome, email, status, obs])


def monta_email(usuario, nome_destinatario, email_destinatario, caminho_pdf: Path):
    import base64

    primeiro_nome = nome_destinatario.split()[0].title()

    msg = MIMEMultipart("related")
    msg["Subject"] = ASSUNTO_EMAIL
    msg["From"] = usuario
    msg["To"] = email_destinatario
    msg["Date"] = formatdate(localtime=True)
    msg["Message-ID"] = make_msgid()

    alt = MIMEMultipart("alternative")
    msg.attach(alt)

    html = TEMPLATE_HTML.format(nome_primeiro=primeiro_nome)
    alt.attach(MIMEText(html, "html", "utf-8"))

    if LOGO1_B64.strip() and not LOGO1_B64.startswith("__"):
        img1 = MIMEImage(base64.b64decode(LOGO1_B64))
        img1.add_header("Content-ID", "<logo1>")
        img1.add_header("Content-Disposition", "inline", filename="logo1.png")
        msg.attach(img1)

    if LOGO2_B64.strip() and not LOGO2_B64.startswith("__"):
        img2 = MIMEImage(base64.b64decode(LOGO2_B64))
        img2.add_header("Content-ID", "<logo2>")
        img2.add_header("Content-Disposition", "inline", filename="logo2.png")
        msg.attach(img2)

    with open(caminho_pdf, "rb") as f:
        anexo = MIMEApplication(f.read(), _subtype="pdf")
    anexo.add_header("Content-Disposition", "attachment", filename=caminho_pdf.name)
    msg.attach(anexo)

    return msg


def salva_em_enviados(imap_host, imap_porta, usuario, senha, msg_bytes, log_fn, pasta_preferida=None):
    """Tenta salvar uma cópia do e-mail já enviado na pasta Enviados via IMAP.
    Retorna o nome da pasta usada com sucesso, ou None se nenhuma funcionou."""
    candidatas = ([pasta_preferida] if pasta_preferida else []) + PASTAS_ENVIADOS_CANDIDATAS
    try:
        imap = imaplib.IMAP4_SSL(imap_host, imap_porta, timeout=20)
        imap.login(usuario, senha)
    except Exception as e:
        log_fn(f"  [aviso] Não consegui logar no IMAP pra salvar em Enviados: {e}")
        return None

    pasta_ok = None
    for pasta in candidatas:
        if not pasta:
            continue
        try:
            typ, _ = imap.select(pasta, readonly=False)
            if typ != "OK":
                continue
            data_imap = imaplib.Time2Internaldate(time.time())
            typ, _ = imap.append(pasta, "\\Seen", data_imap, msg_bytes)
            if typ == "OK":
                pasta_ok = pasta
                break
        except Exception:
            continue

    try:
        imap.logout()
    except Exception:
        pass

    if not pasta_ok:
        log_fn("  [aviso] Não achei a pasta 'Enviados' automaticamente (tente informar o nome exato nas configurações).")
    return pasta_ok


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Envio de Credenciais - CHE")
        self.geometry("760x680")

        self.pasta_pdfs = tk.StringVar()
        self.planilha = tk.StringVar()
        self.coluna_nome = tk.StringVar(value=COLUNA_NOME_PADRAO)
        self.coluna_email = tk.StringVar(value=COLUNA_EMAIL_PADRAO)
        self.smtp_host = tk.StringVar()
        self.smtp_port = tk.StringVar(value="587")
        self.smtp_ssl = tk.BooleanVar(value=False)
        self.smtp_usuario = tk.StringVar()
        self.smtp_senha = tk.StringVar()
        self.imap_host = tk.StringVar()
        self.imap_port = tk.StringVar(value="993")
        self.salvar_enviados = tk.BooleanVar(value=True)
        self.pasta_enviados = tk.StringVar()
        self.log_path = tk.StringVar(value=str(Path.cwd() / "enviados.csv"))

        self._pasta_enviados_confirmada = None
        self._carrega_config()
        self._monta_interface()
        self._matches_atuais = []

    def _carrega_config(self):
        if CONFIG_FILE.exists():
            try:
                cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
                self.pasta_pdfs.set(cfg.get("pasta_pdfs", ""))
                self.planilha.set(cfg.get("planilha", ""))
                self.coluna_nome.set(cfg.get("coluna_nome", COLUNA_NOME_PADRAO))
                self.coluna_email.set(cfg.get("coluna_email", COLUNA_EMAIL_PADRAO))
                self.smtp_host.set(cfg.get("smtp_host", ""))
                self.smtp_port.set(cfg.get("smtp_port", "587"))
                self.smtp_ssl.set(cfg.get("smtp_ssl", False))
                self.smtp_usuario.set(cfg.get("smtp_usuario", ""))
                self.imap_host.set(cfg.get("imap_host", ""))
                self.imap_port.set(cfg.get("imap_port", "993"))
                self.salvar_enviados.set(cfg.get("salvar_enviados", True))
                self.pasta_enviados.set(cfg.get("pasta_enviados", ""))
            except Exception:
                pass

    def _salva_config(self):
        cfg = {
            "pasta_pdfs": self.pasta_pdfs.get(), "planilha": self.planilha.get(),
            "coluna_nome": self.coluna_nome.get(), "coluna_email": self.coluna_email.get(),
            "smtp_host": self.smtp_host.get(), "smtp_port": self.smtp_port.get(),
            "smtp_ssl": self.smtp_ssl.get(), "smtp_usuario": self.smtp_usuario.get(),
            "imap_host": self.imap_host.get(), "imap_port": self.imap_port.get(),
            "salvar_enviados": self.salvar_enviados.get(), "pasta_enviados": self.pasta_enviados.get(),
        }
        CONFIG_FILE.write_text(json.dumps(cfg, indent=2), encoding="utf-8")

    def _monta_interface(self):
        pad = {"padx": 8, "pady": 4}
        frm_arquivos = ttk.LabelFrame(self, text="Arquivos")
        frm_arquivos.pack(fill="x", **pad)
        self._linha_arquivo(frm_arquivos, "Pasta com os PDFs:", self.pasta_pdfs, pasta=True)
        self._linha_arquivo(frm_arquivos, "Planilha de admissão (.xlsx):", self.planilha, pasta=False)

        frm_colunas = ttk.Frame(frm_arquivos)
        frm_colunas.pack(fill="x", padx=8, pady=4)
        ttk.Label(frm_colunas, text="Coluna do nome:").grid(row=0, column=0, sticky="w")
        ttk.Entry(frm_colunas, textvariable=self.coluna_nome, width=25).grid(row=0, column=1, padx=6)
        ttk.Label(frm_colunas, text="Coluna do e-mail:").grid(row=0, column=2, sticky="w")
        ttk.Entry(frm_colunas, textvariable=self.coluna_email, width=25).grid(row=0, column=3, padx=6)

        frm_smtp = ttk.LabelFrame(self, text="Envio (SMTP)")
        frm_smtp.pack(fill="x", **pad)
        l1 = ttk.Frame(frm_smtp); l1.pack(fill="x", padx=8, pady=4)
        ttk.Label(l1, text="Servidor SMTP:", width=16).pack(side="left")
        ttk.Entry(l1, textvariable=self.smtp_host, width=30).pack(side="left", padx=6)
        ttk.Label(l1, text="Porta:").pack(side="left", padx=(12, 0))
        ttk.Entry(l1, textvariable=self.smtp_port, width=6).pack(side="left", padx=6)
        ttk.Checkbutton(l1, text="SSL direto (465)", variable=self.smtp_ssl).pack(side="left", padx=12)

        l2 = ttk.Frame(frm_smtp); l2.pack(fill="x", padx=8, pady=4)
        ttk.Label(l2, text="Usuário (e-mail):", width=16).pack(side="left")
        ttk.Entry(l2, textvariable=self.smtp_usuario, width=30).pack(side="left", padx=6)
        ttk.Label(l2, text="Senha:").pack(side="left", padx=(12, 0))
        ttk.Entry(l2, textvariable=self.smtp_senha, width=20, show="*").pack(side="left", padx=6)

        frm_imap = ttk.LabelFrame(self, text="Salvar cópia em 'Enviados' (IMAP) - opcional")
        frm_imap.pack(fill="x", **pad)
        l3 = ttk.Frame(frm_imap); l3.pack(fill="x", padx=8, pady=4)
        ttk.Checkbutton(l3, text="Salvar cópia em Enviados", variable=self.salvar_enviados).pack(side="left")
        ttk.Label(l3, text="Servidor IMAP:").pack(side="left", padx=(12, 0))
        ttk.Entry(l3, textvariable=self.imap_host, width=26).pack(side="left", padx=6)
        ttk.Label(l3, text="Porta:").pack(side="left", padx=(6, 0))
        ttk.Entry(l3, textvariable=self.imap_port, width=6).pack(side="left", padx=6)
        l4 = ttk.Frame(frm_imap); l4.pack(fill="x", padx=8, pady=4)
        ttk.Label(l4, text="Nome exato da pasta (opcional):").pack(side="left")
        ttk.Entry(l4, textvariable=self.pasta_enviados, width=20).pack(side="left", padx=6)
        ttk.Label(l4, text="(deixe em branco pra tentar detectar automaticamente)", foreground="#888").pack(side="left")

        frm_botoes = ttk.Frame(self)
        frm_botoes.pack(fill="x", **pad)
        ttk.Button(frm_botoes, text="1) Testar (não envia nada)", command=self.testar).pack(side="left", padx=4)
        self.btn_enviar = ttk.Button(frm_botoes, text="2) Enviar de verdade", command=self.enviar)
        self.btn_enviar.pack(side="left", padx=4)
        ttk.Button(frm_botoes, text="Salvar configurações", command=self._salva_config).pack(side="left", padx=4)

        self.progress = ttk.Progressbar(self, mode="determinate")
        self.progress.pack(fill="x", padx=8, pady=(0, 4))

        self.log_box = scrolledtext.ScrolledText(self, height=20)
        self.log_box.pack(fill="both", expand=True, padx=8, pady=8)

    def _linha_arquivo(self, parent, label, var, pasta):
        frm = ttk.Frame(parent)
        frm.pack(fill="x", padx=8, pady=4)
        ttk.Label(frm, text=label, width=26).pack(side="left")
        ttk.Entry(frm, textvariable=var).pack(side="left", fill="x", expand=True, padx=6)

        def escolher():
            caminho = filedialog.askdirectory() if pasta else filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx *.xls")])
            if caminho:
                var.set(caminho)

        ttk.Button(frm, text="Escolher...", command=escolher).pack(side="left")

    def log(self, texto):
        self.log_box.insert("end", texto + "\n")
        self.log_box.see("end")
        self.update_idletasks()

    def _valida_campos_basicos(self):
        if not self.pasta_pdfs.get() or not os.path.isdir(self.pasta_pdfs.get()):
            messagebox.showerror("Erro", "Escolha uma pasta de PDFs válida.")
            return False
        if not self.planilha.get() or not os.path.isfile(self.planilha.get()):
            messagebox.showerror("Erro", "Escolha o arquivo da planilha (.xlsx).")
            return False
        return True

    def testar(self):
        if not self._valida_campos_basicos():
            return
        self.log_box.delete("1.0", "end")
        try:
            mapa = carrega_planilha(self.planilha.get(), self.coluna_nome.get(), self.coluna_email.get())
        except Exception as e:
            messagebox.showerror("Erro na planilha", str(e))
            return

        self.log(f"{len(mapa)} colaboradores na planilha.\n")
        resultados = monta_matches(self.pasta_pdfs.get(), mapa)
        if not resultados:
            self.log("Nenhum PDF encontrado na pasta.")
            return

        ja_enviados = carrega_log(self.log_path.get())
        prontos = []
        for r in resultados:
            pdf_nome = r["pdf"].name
            if r["match"]:
                nome, email = r["match"]
                tag = " (já enviado antes)" if pdf_nome in ja_enviados else ""
                self.log(f"[OK] {pdf_nome}  ->  {nome} <{email}>{tag}")
                if pdf_nome not in ja_enviados:
                    prontos.append((r["pdf"], nome, email))
            else:
                self.log(f"[SEM MATCH] {pdf_nome}   (nome extraído: '{r['nome_extraido']}')")

        self._matches_atuais = prontos
        self.log(f"\n{len(prontos)} e-mail(s) prontos para envio (modo teste - nada foi enviado).")

    def enviar(self):
        if not self._matches_atuais:
            if messagebox.askyesno("Nenhum teste rodado", "Rodar o teste agora?"):
                self.testar()
            return
        if not self.smtp_host.get() or not self.smtp_usuario.get() or not self.smtp_senha.get():
            messagebox.showerror("Erro", "Preencha servidor SMTP, usuário e senha.")
            return
        if not messagebox.askyesno("Confirmar envio", f"Vai enviar {len(self._matches_atuais)} e-mail(s) de verdade. Confirma?"):
            return
        self.btn_enviar.config(state="disabled")
        threading.Thread(target=self._enviar_thread, daemon=True).start()

    def _enviar_thread(self):
        host = self.smtp_host.get()
        try:
            port = int(self.smtp_port.get())
        except ValueError:
            self.log("[ERRO] Porta inválida.")
            self.btn_enviar.config(state="normal")
            return
        usuario, senha = self.smtp_usuario.get(), self.smtp_senha.get()

        self.log(f"\nConectando em {host}:{port} ...")
        try:
            if self.smtp_ssl.get():
                smtp = smtplib.SMTP_SSL(host, port, timeout=20)
            else:
                smtp = smtplib.SMTP(host, port, timeout=20)
                smtp.starttls()
            smtp.login(usuario, senha)
        except Exception as e:
            self.log(f"[ERRO] Não consegui conectar/logar: {e}")
            messagebox.showerror("Erro de conexão", str(e))
            self.btn_enviar.config(state="normal")
            return

        self.log("Conectado! Enviando...\n")
        total = len(self._matches_atuais)
        self.progress.config(maximum=total, value=0)
        enviados_ok = 0

        salvar = self.salvar_enviados.get()
        imap_host = self.imap_host.get().strip()
        try:
            imap_port = int(self.imap_port.get())
        except ValueError:
            imap_port = 993

        for caminho_pdf, nome, email in self._matches_atuais:
            try:
                msg = monta_email(usuario, nome, email, caminho_pdf)
                smtp.sendmail(usuario, [email], msg.as_bytes())
                self.log(f"[ENVIADO] {caminho_pdf.name} -> {email}")
                registra_log(self.log_path.get(), caminho_pdf.name, nome, email, "ENVIADO")
                enviados_ok += 1

                if salvar and imap_host:
                    pasta_usada = salva_em_enviados(
                        imap_host, imap_port, usuario, senha, msg.as_bytes(),
                        self.log, pasta_preferida=self.pasta_enviados.get().strip() or None,
                    )
                    if pasta_usada:
                        if not self.pasta_enviados.get().strip():
                            self.pasta_enviados.set(pasta_usada)
                        self.log(f"  -> cópia salva em '{pasta_usada}'")
            except Exception as e:
                self.log(f"[FALHOU] {caminho_pdf.name} -> {email}: {e}")
                registra_log(self.log_path.get(), caminho_pdf.name, nome, email, "ERRO", str(e))
            self.progress.step(1)
            self.update_idletasks()

        smtp.quit()
        self.log(f"\nConcluído. {enviados_ok}/{total} enviados.")
        self._matches_atuais = []
        self.btn_enviar.config(state="normal")
        messagebox.showinfo("Concluído", f"{enviados_ok}/{total} e-mails enviados.")


if __name__ == "__main__":
    App().mainloop()
